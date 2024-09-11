import os
import re
from openpyxl import load_workbook
import pandas as pd

def intialLaunch():
    fileCheck("spreadsheets", "Please insert files into spreadsheet folder")
    fileCheck("settings.json")
    jsonValidFormat("settings.json")

def fileCheck(name, errorMsg=None):
    if name not in os.listdir():
        if len(name.split(".")) == 1:
            os.mkdir(name)
        else:
            j = open(name, "w")
            j.write("{\n}")
    if errorMsg != None:
        if not os.listdir(name):
            print(errorMsg)
            exit()

def jsonValidFormat(name):
    with open(name) as f:
        lines = [line.rstrip() for line in f]
        if lines[0] != "{" and lines[-1] != "}":
            print(f"{name} file broken, please delete and restart program")
            exit()

def deidentifier(f, stuName):
    workbook = load_workbook(filename="spreadsheets"+os.sep+f)
    sheet = workbook.active
    max_row = sheet.max_row

    for row in range(4, max_row+1):
        sheet[f"A{row}"] = str(stuName)
    os.remove("spreadsheets" + os.sep+f)
    filename = f.split(".")[0]
    newFName = filename + "0000" + ".xlsx"
    workbook.save(filename="spreadsheets" + os.sep+newFName)
    print("File has been deidentified")
    return newFName

def getStudentNameFromDF():
    students = {}
    # for spreadsheets in folder
    for f in os.listdir('spreadsheets'):
        if re.search(r"\.xlsx", f):
            # get cd to find filepath
            cd = os.getcwd()
            # pull student name from spreadsheet
            personalInfo = pd.read_excel("spreadsheets"+os.sep+f, skiprows=2).iat[0, 0].split(" (")
            if len(personalInfo) > 1:
                f = deidentifier(f, personalInfo[0])
            name = personalInfo[0]
            students[name] = pd.read_excel("spreadsheets"+os.sep+f, skiprows=2)
        else:
            print(f"{f} could not be read (must be .xlsx file)")
    return students

def selectSelfName(studs):
    numStud = 0
    userInput = None
    studsSorted = sorted(studs.keys())
    print("Please type the number that corresponds with your name.")

    while userInput == None:
        for stud in studsSorted:
            numStud += 1
            print(f"{numStud}: {stud}")

        try:
            userInput = int(input(""))
        except:
            print("A valid input must be an integer.")

        if userInput not in range(1, numStud+1):
            userInput = None
            numStud = 0
        
    userName = studsSorted[userInput-1]
    return userName